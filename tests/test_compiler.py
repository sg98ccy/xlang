# ============================================================
# tests.test_compiler: compilation correctness tests
# ============================================================

import pytest
from pathlib import Path
from openpyxl import load_workbook
from exlang import compile_xlang_to_xlsx


# ============================================================
# Basic value type tests
# ============================================================

class TestValueTypes:
    """Test that different value types are compiled correctly."""

    def test_integer_values(self, tmp_path):
        """Integer values are stored as int."""
        xlang = """
        <xworkbook>
          <xsheet name="Integers">
            <xcell addr="A1" v="123"/>
            <xcell addr="A2" v="-456"/>
            <xcell addr="A3" v="0"/>
          </xsheet>
        </xworkbook>
        """
        output = tmp_path / "integers.xlsx"
        compile_xlang_to_xlsx(xlang, output)

        wb = load_workbook(output)
        ws = wb["Integers"]

        assert ws["A1"].value == 123
        assert isinstance(ws["A1"].value, int)
        assert ws["A2"].value == -456
        assert ws["A3"].value == 0

    def test_float_values(self, tmp_path):
        """Float values are stored as float."""
        xlang = """
        <xworkbook>
          <xsheet name="Floats">
            <xcell addr="A1" v="123.45"/>
            <xcell addr="A2" v="-0.5"/>
            <xcell addr="A3" v="0.001"/>
          </xsheet>
        </xworkbook>
        """
        output = tmp_path / "floats.xlsx"
        compile_xlang_to_xlsx(xlang, output)

        wb = load_workbook(output)
        ws = wb["Floats"]

        assert ws["A1"].value == 123.45
        assert isinstance(ws["A1"].value, float)
        assert ws["A2"].value == -0.5
        assert ws["A3"].value == 0.001

    def test_string_values(self, tmp_path):
        """String values are preserved."""
        xlang = """
        <xworkbook>
          <xsheet name="Strings">
            <xcell addr="A1" v="Hello"/>
            <xcell addr="A2" v="World"/>
            <xcell addr="A3" v="Region"/>
          </xsheet>
        </xworkbook>
        """
        output = tmp_path / "strings.xlsx"
        compile_xlang_to_xlsx(xlang, output)

        wb = load_workbook(output)
        ws = wb["Strings"]

        assert ws["A1"].value == "Hello"
        assert ws["A2"].value == "World"
        assert ws["A3"].value == "Region"

    def test_formula_values(self, tmp_path):
        """Formulas are preserved as strings starting with =."""
        xlang = """
        <xworkbook>
          <xsheet name="Formulas">
            <xcell addr="A1" v="10"/>
            <xcell addr="A2" v="20"/>
            <xcell addr="A3" v="=SUM(A1:A2)"/>
            <xcell addr="A4" v="=AVERAGE(A1:A2)"/>
          </xsheet>
        </xworkbook>
        """
        output = tmp_path / "formulas.xlsx"
        compile_xlang_to_xlsx(xlang, output)

        wb = load_workbook(output, data_only=False)
        ws = wb["Formulas"]

        assert ws["A1"].value == 10
        assert ws["A2"].value == 20
        assert ws["A3"].value == "=SUM(A1:A2)"
        assert ws["A4"].value == "=AVERAGE(A1:A2)"

    def test_leading_zeros_with_type_hint(self, tmp_path):
        """Leading zeros preserved with t='string'."""
        xlang = """
        <xworkbook>
          <xsheet name="Codes">
            <xcell addr="A1" v="00123" t="string"/>
            <xcell addr="A2" v="007" t="string"/>
          </xsheet>
        </xworkbook>
        """
        output = tmp_path / "codes.xlsx"
        compile_xlang_to_xlsx(xlang, output)

        wb = load_workbook(output)
        ws = wb["Codes"]

        assert ws["A1"].value == "00123"
        assert ws["A2"].value == "007"

    def test_boolean_values(self, tmp_path):
        """Boolean values with type hint."""
        xlang = """
        <xworkbook>
          <xsheet name="Bools">
            <xcell addr="A1" v="TRUE" t="bool"/>
            <xcell addr="A2" v="FALSE" t="bool"/>
            <xcell addr="A3" v="YES" t="bool"/>
            <xcell addr="A4" v="NO" t="bool"/>
          </xsheet>
        </xworkbook>
        """
        output = tmp_path / "bools.xlsx"
        compile_xlang_to_xlsx(xlang, output)

        wb = load_workbook(output)
        ws = wb["Bools"]

        assert ws["A1"].value is True
        assert ws["A2"].value is False
        assert ws["A3"].value is True
        assert ws["A4"].value is False


# ============================================================
# Row-based placement tests
# ============================================================

class TestRowPlacement:
    """Test xrow element placement logic."""

    def test_xrow_default_column(self, tmp_path):
        """xrow defaults to column A."""
        xlang = """
        <xworkbook>
          <xsheet name="Data">
            <xrow r="1"><xv>A</xv><xv>B</xv><xv>C</xv></xrow>
          </xsheet>
        </xworkbook>
        """
        output = tmp_path / "row_default.xlsx"
        compile_xlang_to_xlsx(xlang, output)

        wb = load_workbook(output)
        ws = wb["Data"]

        assert ws["A1"].value == "A"
        assert ws["B1"].value == "B"
        assert ws["C1"].value == "C"

    def test_xrow_custom_column(self, tmp_path):
        """xrow with c attribute starts at specified column."""
        xlang = """
        <xworkbook>
          <xsheet name="Data">
            <xrow r="1" c="C"><xv>X</xv><xv>Y</xv><xv>Z</xv></xrow>
          </xsheet>
        </xworkbook>
        """
        output = tmp_path / "row_custom.xlsx"
        compile_xlang_to_xlsx(xlang, output)

        wb = load_workbook(output)
        ws = wb["Data"]

        assert ws["C1"].value == "X"
        assert ws["D1"].value == "Y"
        assert ws["E1"].value == "Z"

    def test_multiple_rows(self, tmp_path):
        """Multiple xrow elements work correctly."""
        xlang = """
        <xworkbook>
          <xsheet name="Data">
            <xrow r="1"><xv>Region</xv><xv>Sales</xv></xrow>
            <xrow r="2"><xv>North</xv><xv>120000</xv></xrow>
            <xrow r="3"><xv>South</xv><xv>98000</xv></xrow>
          </xsheet>
        </xworkbook>
        """
        output = tmp_path / "multi_rows.xlsx"
        compile_xlang_to_xlsx(xlang, output)

        wb = load_workbook(output)
        ws = wb["Data"]

        assert ws["A1"].value == "Region"
        assert ws["B1"].value == "Sales"
        assert ws["A2"].value == "North"
        assert ws["B2"].value == 120000
        assert ws["A3"].value == "South"
        assert ws["B3"].value == 98000


# ============================================================
# Multi-sheet tests
# ============================================================

class TestMultiSheet:
    """Test workbooks with multiple sheets."""

    def test_two_sheets(self, tmp_path):
        """Two sheets are created correctly."""
        xlang = """
        <xworkbook>
          <xsheet name="Data">
            <xcell addr="A1" v="Data Sheet"/>
          </xsheet>
          <xsheet name="Summary">
            <xcell addr="A1" v="Summary Sheet"/>
          </xsheet>
        </xworkbook>
        """
        output = tmp_path / "two_sheets.xlsx"
        compile_xlang_to_xlsx(xlang, output)

        wb = load_workbook(output)

        assert "Data" in wb.sheetnames
        assert "Summary" in wb.sheetnames
        assert wb["Data"]["A1"].value == "Data Sheet"
        assert wb["Summary"]["A1"].value == "Summary Sheet"

    def test_cross_sheet_formula(self, tmp_path):
        """Formulas can reference other sheets."""
        xlang = """
        <xworkbook>
          <xsheet name="Data">
            <xcell addr="A1" v="100"/>
            <xcell addr="A2" v="200"/>
          </xsheet>
          <xsheet name="Summary">
            <xcell addr="A1" v="=Data!A1+Data!A2"/>
          </xsheet>
        </xworkbook>
        """
        output = tmp_path / "cross_sheet.xlsx"
        compile_xlang_to_xlsx(xlang, output)

        wb = load_workbook(output, data_only=False)

        assert wb["Data"]["A1"].value == 100
        assert wb["Data"]["A2"].value == 200
        assert wb["Summary"]["A1"].value == "=Data!A1+Data!A2"


# ============================================================
# Mixed xrow and xcell tests
# ============================================================

class TestMixedPlacement:
    """Test documents with both xrow and xcell."""

    def test_xrow_and_xcell_combined(self, tmp_path):
        """xrow and xcell can coexist."""
        xlang = """
        <xworkbook>
          <xsheet name="Mixed">
            <xrow r="1"><xv>Header1</xv><xv>Header2</xv></xrow>
            <xrow r="2"><xv>Value1</xv><xv>Value2</xv></xrow>
            <xcell addr="A4" v="Total"/>
            <xcell addr="B4" v="=SUM(B2:B3)"/>
          </xsheet>
        </xworkbook>
        """
        output = tmp_path / "mixed.xlsx"
        compile_xlang_to_xlsx(xlang, output)

        wb = load_workbook(output, data_only=False)
        ws = wb["Mixed"]

        assert ws["A1"].value == "Header1"
        assert ws["B1"].value == "Header2"
        assert ws["A2"].value == "Value1"
        assert ws["B2"].value == "Value2"
        assert ws["A4"].value == "Total"
        assert ws["B4"].value == "=SUM(B2:B3)"
