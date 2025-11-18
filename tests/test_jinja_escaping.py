"""
Test Jinja2 auto-escaping for EXLang formulas.

Jinja2 preprocessing is now AUTOMATIC in EXLang compiler.
This demonstrates how the template engine handles XML escaping automatically.
"""

from exlang import compile_xlang_to_xlsx
from openpyxl import load_workbook
from pathlib import Path


def test_jinja_basic_formula(tmp_path):
    """Test basic formula with < and quotes using Jinja2 variables."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xcell addr="A1" v="{{ formula }}"/>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "test.xlsx"
    
    # Pass formula as variable - Jinja2 auto-escapes it automatically
    compile_xlang_to_xlsx(
        xlang, 
        output,
        formula='=IF(B1<100,"Low","High")'
    )
    
    # Verify the file was created and formula is correct
    wb = load_workbook(output)
    ws = wb.active
    assert ws['A1'].value == '=IF(B1<100,"Low","High")'


def test_jinja_multiple_formulas(tmp_path):
    """Test multiple formulas with different operators."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xcell addr="A1" v="{{ less_than }}"/>
        <xcell addr="A2" v="{{ greater_than }}"/>
        <xcell addr="A3" v="{{ not_equal }}"/>
        <xcell addr="A4" v="{{ text_ampersand }}"/>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "test.xlsx"
    
    compile_xlang_to_xlsx(
        xlang,
        output,
        less_than='=IF(B1<100,"Low","High")',
        greater_than='=IF(B2>100,"High","Low")',
        not_equal='=IF(B3<>0,"Active","Inactive")',
        text_ampersand='Sales & Marketing'
    )
    
    wb = load_workbook(output)
    ws = wb.active
    assert ws['A1'].value == '=IF(B1<100,"Low","High")'
    assert ws['A2'].value == '=IF(B2>100,"High","Low")'
    assert ws['A3'].value == '=IF(B3<>0,"Active","Inactive")'
    assert ws['A4'].value == 'Sales & Marketing'


def test_jinja_loop_formulas(tmp_path):
    """Test Jinja2 loops for generating repetitive formulas."""
    xlang = """
    <xworkbook>
      <xsheet name="Inventory">
        {% for row in rows %}
        <xcell addr="K{{ row }}" v="{{ formula_template.replace('ROW', row|string) }}"/>
        {% endfor %}
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "test.xlsx"
    
    compile_xlang_to_xlsx(
        xlang,
        output,
        rows=[4, 5, 6],
        formula_template='=IF(JROW<100,"REORDER","OK")'
    )
    
    wb = load_workbook(output)
    ws = wb.active
    assert ws['K4'].value == '=IF(J4<100,"REORDER","OK")'
    assert ws['K5'].value == '=IF(J5<100,"REORDER","OK")'
    assert ws['K6'].value == '=IF(J6<100,"REORDER","OK")'


def test_jinja_complex_nested_formula(tmp_path):
    """Test complex nested formulas with multiple operators."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xcell addr="A1" v="{{ complex_formula }}"/>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "test.xlsx"
    
    compile_xlang_to_xlsx(
        xlang,
        output,
        complex_formula='=IF(AND(B1>50,B1<100),"Medium",IF(B1>=100,"High","Low"))'
    )
    
    wb = load_workbook(output)
    ws = wb.active
    assert ws['A1'].value == '=IF(AND(B1>50,B1<100),"Medium",IF(B1>=100,"High","Low"))'


def test_plain_xml_with_manual_escaping(tmp_path):
    """Test that plain XML with manual escaping still works (backward compatibility)."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xcell addr="A1" v="=IF(B1&lt;100,&quot;Low&quot;,&quot;High&quot;)"/>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "test.xlsx"
    
    # Manual escaping still works with automatic Jinja2 preprocessing
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    ws = wb.active
    assert ws['A1'].value == '=IF(B1<100,"Low","High")'
