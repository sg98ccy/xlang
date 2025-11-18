"""
Tests for xrange feature.
"""

import pytest
from pathlib import Path
from openpyxl import load_workbook

from exlang import compile_xlang_to_xlsx


# ============================================================
# Tests: xrange basic functionality
# ============================================================

def test_xrange_single_cell(tmp_path):
    """Test xrange with single cell (from = to)."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrange from="B2" to="B2" fill="100"/>
      </xsheet>
    </xworkbook>
    """.strip()
    
    output = tmp_path / "test.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    ws = wb["Test"]
    assert ws["B2"].value == 100


def test_xrange_row(tmp_path):
    """Test xrange filling a row."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrange from="A1" to="E1" fill="Header"/>
      </xsheet>
    </xworkbook>
    """.strip()
    
    output = tmp_path / "test.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    ws = wb["Test"]
    
    for col in range(1, 6):  # A to E (columns 1-5)
        assert ws.cell(row=1, column=col).value == "Header"


def test_xrange_column(tmp_path):
    """Test xrange filling a column."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrange from="C2" to="C10" fill="0"/>
      </xsheet>
    </xworkbook>
    """.strip()
    
    output = tmp_path / "test.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    ws = wb["Test"]
    
    for row in range(2, 11):  # Rows 2-10
        assert ws.cell(row=row, column=3).value == 0


def test_xrange_rectangular_area(tmp_path):
    """Test xrange filling a rectangular area."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrange from="B2" to="D4" fill="X"/>
      </xsheet>
    </xworkbook>
    """.strip()
    
    output = tmp_path / "test.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    ws = wb["Test"]
    
    # Check 3x3 area is filled
    for row in range(2, 5):  # Rows 2-4
        for col in range(2, 5):  # Cols B-D (2-4)
            assert ws.cell(row=row, column=col).value == "X"
    
    # Check surrounding cells are empty
    assert ws["A1"].value is None
    assert ws["E5"].value is None


def test_xrange_large_area(tmp_path):
    """Test xrange with large area (B2:B50)."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrange from="B2" to="B50" fill="0"/>
      </xsheet>
    </xworkbook>
    """.strip()
    
    output = tmp_path / "test.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    ws = wb["Test"]
    
    for row in range(2, 51):  # Rows 2-50
        assert ws.cell(row=row, column=2).value == 0


# ============================================================
# Tests: xrange type inference
# ============================================================

def test_xrange_integer_fill(tmp_path):
    """Test xrange with integer fill value."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrange from="A1" to="A3" fill="42"/>
      </xsheet>
    </xworkbook>
    """.strip()
    
    output = tmp_path / "test.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    ws = wb["Test"]
    
    for row in range(1, 4):
        assert ws.cell(row=row, column=1).value == 42
        assert isinstance(ws.cell(row=row, column=1).value, int)


def test_xrange_float_fill(tmp_path):
    """Test xrange with float fill value."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrange from="A1" to="A3" fill="3.14"/>
      </xsheet>
    </xworkbook>
    """.strip()
    
    output = tmp_path / "test.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    ws = wb["Test"]
    
    for row in range(1, 4):
        assert ws.cell(row=row, column=1).value == 3.14
        assert isinstance(ws.cell(row=row, column=1).value, float)


def test_xrange_string_fill(tmp_path):
    """Test xrange with string fill value."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrange from="A1" to="A3" fill="N/A"/>
      </xsheet>
    </xworkbook>
    """.strip()
    
    output = tmp_path / "test.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    ws = wb["Test"]
    
    for row in range(1, 4):
        assert ws.cell(row=row, column=1).value == "N/A"


def test_xrange_formula_fill(tmp_path):
    """Test xrange with formula fill value."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrange from="A1" to="A3" fill="=SUM(B1:C1)"/>
      </xsheet>
    </xworkbook>
    """.strip()
    
    output = tmp_path / "test.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    ws = wb["Test"]
    
    for row in range(1, 4):
        # Note: Formula is the same for all cells - this is expected behaviour
        assert ws.cell(row=row, column=1).value == "=SUM(B1:C1)"


# ============================================================
# Tests: xrange with type hints
# ============================================================

def test_xrange_type_hint_string(tmp_path):
    """Test xrange with t='string' preserves leading zeros."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrange from="A1" to="A3" fill="00123" t="string"/>
      </xsheet>
    </xworkbook>
    """.strip()
    
    output = tmp_path / "test.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    ws = wb["Test"]
    
    for row in range(1, 4):
        assert ws.cell(row=row, column=1).value == "00123"


def test_xrange_type_hint_number(tmp_path):
    """Test xrange with t='number' forces numeric conversion."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrange from="A1" to="A3" fill="42" t="number"/>
      </xsheet>
    </xworkbook>
    """.strip()
    
    output = tmp_path / "test.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    ws = wb["Test"]
    
    for row in range(1, 4):
        assert ws.cell(row=row, column=1).value == 42
        assert isinstance(ws.cell(row=row, column=1).value, int)


# ============================================================
# Tests: xrange combined with other tags
# ============================================================

def test_xrange_with_xrow(tmp_path):
    """Test xrange combined with xrow."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrow r="1" c="A">
          <xv>Name</xv>
          <xv>Value</xv>
        </xrow>
        <xrange from="A2" to="A10" fill="Item"/>
        <xrange from="B2" to="B10" fill="0"/>
      </xsheet>
    </xworkbook>
    """.strip()
    
    output = tmp_path / "test.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    ws = wb["Test"]
    
    # Check headers
    assert ws["A1"].value == "Name"
    assert ws["B1"].value == "Value"
    
    # Check ranges
    for row in range(2, 11):
        assert ws.cell(row=row, column=1).value == "Item"
        assert ws.cell(row=row, column=2).value == 0


def test_xrange_with_xcell(tmp_path):
    """Test xrange combined with xcell (xcell should overwrite)."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrange from="A1" to="A5" fill="Default"/>
        <xcell addr="A3" v="Override"/>
      </xsheet>
    </xworkbook>
    """.strip()
    
    output = tmp_path / "test.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    ws = wb["Test"]
    
    assert ws["A1"].value == "Default"
    assert ws["A2"].value == "Default"
    assert ws["A3"].value == "Override"  # Overwritten by xcell
    assert ws["A4"].value == "Default"
    assert ws["A5"].value == "Default"


def test_multiple_xranges(tmp_path):
    """Test multiple xrange tags in same sheet."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrange from="A1" to="A5" fill="1"/>
        <xrange from="B1" to="B5" fill="2"/>
        <xrange from="C1" to="C5" fill="3"/>
      </xsheet>
    </xworkbook>
    """.strip()
    
    output = tmp_path / "test.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    ws = wb["Test"]
    
    for row in range(1, 6):
        assert ws.cell(row=row, column=1).value == 1
        assert ws.cell(row=row, column=2).value == 2
        assert ws.cell(row=row, column=3).value == 3


# ============================================================
# Tests: xrange validation errors
# ============================================================

def test_xrange_missing_from():
    """Test xrange without 'from' attribute fails validation."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrange to="B10" fill="0"/>
      </xsheet>
    </xworkbook>
    """.strip()
    
    with pytest.raises(ValueError, match="missing required attribute 'from'"):
        compile_xlang_to_xlsx(xlang, "output.xlsx")


def test_xrange_missing_to():
    """Test xrange without 'to' attribute fails validation."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrange from="B2" fill="0"/>
      </xsheet>
    </xworkbook>
    """.strip()
    
    with pytest.raises(ValueError, match="missing required attribute 'to'"):
        compile_xlang_to_xlsx(xlang, "output.xlsx")


def test_xrange_missing_fill():
    """Test xrange without 'fill' attribute fails validation."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrange from="B2" to="B10"/>
      </xsheet>
    </xworkbook>
    """.strip()
    
    with pytest.raises(ValueError, match="missing required attribute 'fill'"):
        compile_xlang_to_xlsx(xlang, "output.xlsx")


def test_xrange_invalid_type_hint():
    """Test xrange with invalid type hint fails validation."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrange from="A1" to="A3" fill="0" t="invalid"/>
      </xsheet>
    </xworkbook>
    """.strip()
    
    with pytest.raises(ValueError, match="invalid type hint"):
        compile_xlang_to_xlsx(xlang, "output.xlsx")


def test_xrange_invalid_from_address():
    """Test xrange with invalid 'from' address fails."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrange from="INVALID" to="B10" fill="0"/>
      </xsheet>
    </xworkbook>
    """.strip()
    
    with pytest.raises(ValueError, match="Invalid cell address"):
        compile_xlang_to_xlsx(xlang, "output.xlsx")


def test_xrange_invalid_to_address():
    """Test xrange with invalid 'to' address fails."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrange from="B2" to="999" fill="0"/>
      </xsheet>
    </xworkbook>
    """.strip()
    
    with pytest.raises(ValueError, match="Invalid cell address"):
        compile_xlang_to_xlsx(xlang, "output.xlsx")


def test_xrange_from_after_to():
    """Test xrange where 'from' is after 'to' fails."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrange from="B10" to="B2" fill="0"/>
      </xsheet>
    </xworkbook>
    """.strip()
    
    with pytest.raises(ValueError, match="'from'.*must be before or equal to 'to'"):
        compile_xlang_to_xlsx(xlang, "output.xlsx")
