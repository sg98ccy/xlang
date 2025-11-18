# ============================================================
# tests.test_merge_style: Tests for xmerge and xstyle tags
# ============================================================

from pathlib import Path
import pytest
from openpyxl import load_workbook

from exlang import compile_xlang_to_xlsx, validate_xlang_minimal
from xml.etree import ElementTree as ET


# ============================================================
# Basic xmerge Tests
# ============================================================

def test_xmerge_simple_horizontal(tmp_path):
    """Simple horizontal merge A1:B1."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xcell addr="A1" v="Merged Title"/>
        <xmerge addr="A1:B1"/>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "merge_horizontal.xlsx"
    compile_xlang_to_xlsx(xlang, output)

    wb = load_workbook(output)
    ws = wb["Test"]
    
    # Check that cells are merged
    assert "A1:B1" in [str(mc) for mc in ws.merged_cells.ranges]
    assert ws["A1"].value == "Merged Title"


def test_xmerge_vertical(tmp_path):
    """Vertical merge A1:A3."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xcell addr="A1" v="Merged Column"/>
        <xmerge addr="A1:A3"/>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "merge_vertical.xlsx"
    compile_xlang_to_xlsx(xlang, output)

    wb = load_workbook(output)
    ws = wb["Test"]
    
    assert "A1:A3" in [str(mc) for mc in ws.merged_cells.ranges]
    assert ws["A1"].value == "Merged Column"


def test_xmerge_rectangular(tmp_path):
    """Rectangular merge A1:C3."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xcell addr="A1" v="Large Merge"/>
        <xmerge addr="A1:C3"/>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "merge_rectangular.xlsx"
    compile_xlang_to_xlsx(xlang, output)

    wb = load_workbook(output)
    ws = wb["Test"]
    
    assert "A1:C3" in [str(mc) for mc in ws.merged_cells.ranges]


def test_xmerge_multiple_ranges(tmp_path):
    """Multiple merges in same sheet."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xcell addr="A1" v="Title 1"/>
        <xcell addr="C1" v="Title 2"/>
        <xmerge addr="A1:B1"/>
        <xmerge addr="C1:D1"/>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "merge_multiple.xlsx"
    compile_xlang_to_xlsx(xlang, output)

    wb = load_workbook(output)
    ws = wb["Test"]
    
    merged_ranges = [str(mc) for mc in ws.merged_cells.ranges]
    assert "A1:B1" in merged_ranges
    assert "C1:D1" in merged_ranges


# ============================================================
# Basic xstyle Tests
# ============================================================

def test_xstyle_bold_single_cell(tmp_path):
    """Apply bold to single cell."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xcell addr="A1" v="Bold Text"/>
        <xstyle addr="A1" bold="true"/>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "style_bold.xlsx"
    compile_xlang_to_xlsx(xlang, output)

    wb = load_workbook(output)
    ws = wb["Test"]
    
    assert ws["A1"].value == "Bold Text"
    assert ws["A1"].font.bold is True


def test_xstyle_italic_single_cell(tmp_path):
    """Apply italic to single cell."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xcell addr="A1" v="Italic Text"/>
        <xstyle addr="A1" italic="true"/>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "style_italic.xlsx"
    compile_xlang_to_xlsx(xlang, output)

    wb = load_workbook(output)
    ws = wb["Test"]
    
    assert ws["A1"].font.italic is True


def test_xstyle_underline_single_cell(tmp_path):
    """Apply underline to single cell."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xcell addr="A1" v="Underlined Text"/>
        <xstyle addr="A1" underline="true"/>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "style_underline.xlsx"
    compile_xlang_to_xlsx(xlang, output)

    wb = load_workbook(output)
    ws = wb["Test"]
    
    assert ws["A1"].font.underline == "single"


def test_xstyle_multiple_attributes(tmp_path):
    """Apply bold, italic, and underline together."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xcell addr="A1" v="Formatted Text"/>
        <xstyle addr="A1" bold="true" italic="true" underline="true"/>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "style_multiple.xlsx"
    compile_xlang_to_xlsx(xlang, output)

    wb = load_workbook(output)
    ws = wb["Test"]
    
    assert ws["A1"].font.bold is True
    assert ws["A1"].font.italic is True
    assert ws["A1"].font.underline == "single"


def test_xstyle_false_values(tmp_path):
    """Explicitly set bold=false."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xcell addr="A1" v="Normal Text"/>
        <xstyle addr="A1" bold="false"/>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "style_false.xlsx"
    compile_xlang_to_xlsx(xlang, output)

    wb = load_workbook(output)
    ws = wb["Test"]
    
    assert ws["A1"].font.bold is False


def test_xstyle_range(tmp_path):
    """Apply style to range of cells."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrow r="1" c="A"><xv>Header 1</xv><xv>Header 2</xv><xv>Header 3</xv></xrow>
        <xstyle addr="A1:C1" bold="true"/>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "style_range.xlsx"
    compile_xlang_to_xlsx(xlang, output)

    wb = load_workbook(output)
    ws = wb["Test"]
    
    assert ws["A1"].font.bold is True
    assert ws["B1"].font.bold is True
    assert ws["C1"].font.bold is True


def test_xstyle_large_range(tmp_path):
    """Apply style to larger range."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrange from="A1" to="C3" fill="Data"/>
        <xstyle addr="A1:C3" italic="true"/>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "style_large_range.xlsx"
    compile_xlang_to_xlsx(xlang, output)

    wb = load_workbook(output)
    ws = wb["Test"]
    
    # Check all cells in range
    for row in range(1, 4):
        for col in ["A", "B", "C"]:
            addr = f"{col}{row}"
            assert ws[addr].font.italic is True


# ============================================================
# Integration Tests
# ============================================================

def test_merge_and_style_combined(tmp_path):
    """Merge cells and apply styling."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xcell addr="A1" v="Bold Merged Title"/>
        <xmerge addr="A1:D1"/>
        <xstyle addr="A1" bold="true"/>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "merge_and_style.xlsx"
    compile_xlang_to_xlsx(xlang, output)

    wb = load_workbook(output)
    ws = wb["Test"]
    
    # Check merge
    assert "A1:D1" in [str(mc) for mc in ws.merged_cells.ranges]
    # Check styling
    assert ws["A1"].value == "Bold Merged Title"
    assert ws["A1"].font.bold is True


def test_xstyle_with_xrepeat(tmp_path):
    """Style cells generated by xrepeat."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrepeat times="3" r="1" c="A">
          <xv>Month {{i}}</xv>
        </xrepeat>
        <xstyle addr="A1:A3" bold="true"/>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "style_with_repeat.xlsx"
    compile_xlang_to_xlsx(xlang, output)

    wb = load_workbook(output)
    ws = wb["Test"]
    
    # Check values
    assert ws["A1"].value == "Month 1"
    assert ws["A2"].value == "Month 2"
    assert ws["A3"].value == "Month 3"
    
    # Check styling
    assert ws["A1"].font.bold is True
    assert ws["A2"].font.bold is True
    assert ws["A3"].font.bold is True


def test_realistic_table_header(tmp_path):
    """Realistic use case: styled and merged header."""
    xlang = """
    <xworkbook>
      <xsheet name="Budget">
        <xcell addr="A1" v="2024 Monthly Budget"/>
        <xmerge addr="A1:D1"/>
        <xstyle addr="A1" bold="true"/>
        
        <xrow r="2" c="A"><xv>Month</xv><xv>Income</xv><xv>Expenses</xv><xv>Savings</xv></xrow>
        <xstyle addr="A2:D2" bold="true" italic="true"/>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "realistic_table.xlsx"
    compile_xlang_to_xlsx(xlang, output)

    wb = load_workbook(output)
    ws = wb["Budget"]
    
    # Check merged title
    assert "A1:D1" in [str(mc) for mc in ws.merged_cells.ranges]
    assert ws["A1"].value == "2024 Monthly Budget"
    assert ws["A1"].font.bold is True
    
    # Check header row
    assert ws["A2"].value == "Month"
    assert ws["B2"].value == "Income"
    assert ws["C2"].value == "Expenses"
    assert ws["D2"].value == "Savings"
    
    # Check header styling
    for col in ["A", "B", "C", "D"]:
        assert ws[f"{col}2"].font.bold is True
        assert ws[f"{col}2"].font.italic is True


# ============================================================
# Validation Error Tests
# ============================================================

def test_xmerge_missing_addr():
    """xmerge without addr attribute should fail validation."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xmerge/>
      </xsheet>
    </xworkbook>
    """
    root = ET.fromstring(xlang)
    errors = validate_xlang_minimal(root)
    assert any("xmerge missing required attribute 'addr'" in e for e in errors)


def test_xmerge_invalid_format():
    """xmerge with invalid range format should fail validation."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xmerge addr="A1"/>
      </xsheet>
    </xworkbook>
    """
    root = ET.fromstring(xlang)
    errors = validate_xlang_minimal(root)
    assert any("must be a range" in e for e in errors)


def test_xstyle_missing_addr():
    """xstyle without addr attribute should fail validation."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xstyle bold="true"/>
      </xsheet>
    </xworkbook>
    """
    root = ET.fromstring(xlang)
    errors = validate_xlang_minimal(root)
    assert any("xstyle missing required attribute 'addr'" in e for e in errors)


def test_xstyle_invalid_bool():
    """xstyle with invalid boolean value should fail validation."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xstyle addr="A1" bold="yes"/>
      </xsheet>
    </xworkbook>
    """
    root = ET.fromstring(xlang)
    errors = validate_xlang_minimal(root)
    assert any("invalid bold='yes'" in e for e in errors)


def test_xstyle_multiple_invalid_bools():
    """xstyle with multiple invalid boolean values."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xstyle addr="A1" bold="1" italic="yes" underline="no"/>
      </xsheet>
    </xworkbook>
    """
    root = ET.fromstring(xlang)
    errors = validate_xlang_minimal(root)
    assert any("invalid bold='1'" in e for e in errors)
    assert any("invalid italic='yes'" in e for e in errors)
    assert any("invalid underline='no'" in e for e in errors)


# ============================================================
# Edge Cases
# ============================================================

def test_xstyle_no_attributes_except_addr(tmp_path):
    """xstyle with only addr (no formatting attributes) should work."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xcell addr="A1" v="Normal"/>
        <xstyle addr="A1"/>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "style_no_attrs.xlsx"
    compile_xlang_to_xlsx(xlang, output)

    wb = load_workbook(output)
    ws = wb["Test"]
    
    # Should apply default font (bold=False, italic=False, underline=None)
    assert ws["A1"].font.bold is False
    assert ws["A1"].font.italic is False


def test_multiple_xstyle_on_same_cell(tmp_path):
    """Multiple xstyle tags on same cell (last write wins)."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xcell addr="A1" v="Test"/>
        <xstyle addr="A1" bold="true"/>
        <xstyle addr="A1" italic="true"/>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "style_multiple_tags.xlsx"
    compile_xlang_to_xlsx(xlang, output)

    wb = load_workbook(output)
    ws = wb["Test"]
    
    # Last xstyle wins, so only italic should be true
    # (bold from first xstyle is overwritten)
    assert ws["A1"].font.bold is False
    assert ws["A1"].font.italic is True
