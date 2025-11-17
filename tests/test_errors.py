# ============================================================
# tests.test_errors: error handling and robustness tests
# ============================================================

import pytest
from pathlib import Path
from xml.etree import ElementTree as ET
from exlang import compile_xlang_to_xlsx, validate_xlang_minimal


# ============================================================
# Validation error tests
# ============================================================

class TestValidationErrors:
    """Test that validation errors prevent compilation."""

    def test_invalid_root_prevents_compilation(self, tmp_path):
        """Invalid root tag raises ValueError."""
        xlang = "<workbook><xsheet name='Test'></xsheet></workbook>"
        output = tmp_path / "invalid.xlsx"

        with pytest.raises(ValueError, match="Invalid XLang"):
            compile_xlang_to_xlsx(xlang, output)

    def test_missing_sheet_name_prevents_compilation(self, tmp_path):
        """Missing sheet name raises ValueError."""
        xlang = "<xworkbook><xsheet></xsheet></xworkbook>"
        output = tmp_path / "invalid.xlsx"

        with pytest.raises(ValueError, match="Invalid XLang"):
            compile_xlang_to_xlsx(xlang, output)

    def test_missing_xrow_r_prevents_compilation(self, tmp_path):
        """Missing xrow r attribute raises ValueError."""
        xlang = """
        <xworkbook>
          <xsheet name="Data">
            <xrow><xv>A</xv></xrow>
          </xsheet>
        </xworkbook>
        """
        output = tmp_path / "invalid.xlsx"

        with pytest.raises(ValueError, match="Invalid XLang"):
            compile_xlang_to_xlsx(xlang, output)

    def test_missing_xcell_addr_prevents_compilation(self, tmp_path):
        """Missing xcell addr raises ValueError."""
        xlang = """
        <xworkbook>
          <xsheet name="Data">
            <xcell v="123"/>
          </xsheet>
        </xworkbook>
        """
        output = tmp_path / "invalid.xlsx"

        with pytest.raises(ValueError, match="Invalid XLang"):
            compile_xlang_to_xlsx(xlang, output)

    def test_invalid_type_hint_prevents_compilation(self, tmp_path):
        """Invalid type hint raises ValueError."""
        xlang = """
        <xworkbook>
          <xsheet name="Data">
            <xcell addr="A1" v="123" t="invalid"/>
          </xsheet>
        </xworkbook>
        """
        output = tmp_path / "invalid.xlsx"

        with pytest.raises(ValueError, match="Invalid XLang"):
            compile_xlang_to_xlsx(xlang, output)


# ============================================================
# XML parsing error tests
# ============================================================

class TestXMLParsingErrors:
    """Test that malformed XML is caught early."""

    def test_malformed_xml(self, tmp_path):
        """Malformed XML raises ET.ParseError."""
        xlang = "<xworkbook><xsheet name='Test'>"  # Missing closing tag
        output = tmp_path / "invalid.xlsx"

        with pytest.raises(ET.ParseError):
            compile_xlang_to_xlsx(xlang, output)

    def test_empty_string(self, tmp_path):
        """Empty string raises ParseError."""
        xlang = ""
        output = tmp_path / "invalid.xlsx"

        with pytest.raises(ET.ParseError):
            compile_xlang_to_xlsx(xlang, output)

    def test_whitespace_only(self, tmp_path):
        """Whitespace-only string raises ParseError."""
        xlang = "   \n\n   "
        output = tmp_path / "invalid.xlsx"

        with pytest.raises(ET.ParseError):
            compile_xlang_to_xlsx(xlang, output)


# ============================================================
# Runtime error tests
# ============================================================

class TestRuntimeErrors:
    """Test runtime errors during compilation."""

    def test_invalid_column_letter_in_xrow(self, tmp_path):
        """Invalid column letter in xrow c attribute raises error."""
        xlang = """
        <xworkbook>
          <xsheet name="Data">
            <xrow r="1" c="A1"><xv>Test</xv></xrow>
          </xsheet>
        </xworkbook>
        """
        output = tmp_path / "invalid.xlsx"

        with pytest.raises(ValueError, match="Invalid column letter"):
            compile_xlang_to_xlsx(xlang, output)

    def test_non_numeric_row_index(self, tmp_path):
        """Non-numeric row index raises ValueError."""
        xlang = """
        <xworkbook>
          <xsheet name="Data">
            <xrow r="one"><xv>Test</xv></xrow>
          </xsheet>
        </xworkbook>
        """
        output = tmp_path / "invalid.xlsx"

        with pytest.raises(ValueError):
            compile_xlang_to_xlsx(xlang, output)


# ============================================================
# File system error tests
# ============================================================

class TestFileSystemErrors:
    """Test file system related errors."""

    def test_output_directory_created(self, tmp_path):
        """Output directory is created if it doesn't exist."""
        xlang = """
        <xworkbook>
          <xsheet name="Test">
            <xcell addr="A1" v="Hello"/>
          </xsheet>
        </xworkbook>
        """
        # Use a nested path that doesn't exist
        output = tmp_path / "nested" / "subdir" / "test.xlsx"

        # Should not raise error, directory should be created
        compile_xlang_to_xlsx(xlang, output)

        assert output.exists()
        assert output.is_file()


# ============================================================
# Edge case error handling
# ============================================================

class TestEdgeCaseErrors:
    """Test edge cases that might cause errors."""

    def test_empty_xv_element(self, tmp_path):
        """Empty xv elements are handled (empty string)."""
        xlang = """
        <xworkbook>
          <xsheet name="Data">
            <xrow r="1"><xv></xv><xv>B</xv></xrow>
          </xsheet>
        </xworkbook>
        """
        output = tmp_path / "empty_xv.xlsx"
        compile_xlang_to_xlsx(xlang, output)

        from openpyxl import load_workbook
        wb = load_workbook(output)
        ws = wb["Data"]

        # Empty xv elements return None in openpyxl
        assert ws["A1"].value is None
        assert ws["B1"].value == "B"

    def test_special_characters_in_sheet_name(self, tmp_path):
        """Sheet names with special characters work."""
        xlang = """
        <xworkbook>
          <xsheet name="Data-2025">
            <xcell addr="A1" v="Test"/>
          </xsheet>
        </xworkbook>
        """
        output = tmp_path / "special_chars.xlsx"
        compile_xlang_to_xlsx(xlang, output)

        from openpyxl import load_workbook
        wb = load_workbook(output)
        assert "Data-2025" in wb.sheetnames

    def test_unicode_in_values(self, tmp_path):
        """Unicode characters in values are preserved."""
        xlang = """
        <xworkbook>
          <xsheet name="Unicode">
            <xcell addr="A1" v="Hello 世界"/>
            <xcell addr="A2" v="Café"/>
            <xcell addr="A3" v="🚀"/>
          </xsheet>
        </xworkbook>
        """
        output = tmp_path / "unicode.xlsx"
        compile_xlang_to_xlsx(xlang, output)

        from openpyxl import load_workbook
        wb = load_workbook(output)
        ws = wb["Unicode"]

        assert ws["A1"].value == "Hello 世界"
        assert ws["A2"].value == "Café"
        assert ws["A3"].value == "🚀"

    def test_very_long_string(self, tmp_path):
        """Very long strings are handled."""
        long_string = "A" * 10000
        xlang = f"""
        <xworkbook>
          <xsheet name="Long">
            <xcell addr="A1" v="{long_string}"/>
          </xsheet>
        </xworkbook>
        """
        output = tmp_path / "long_string.xlsx"
        compile_xlang_to_xlsx(xlang, output)

        from openpyxl import load_workbook
        wb = load_workbook(output)
        ws = wb["Long"]

        assert ws["A1"].value == long_string
        assert len(ws["A1"].value) == 10000
