# ============================================================
# tests.test_roundtrip: end-to-end semantic preservation tests
# ============================================================

import pytest
from pathlib import Path
from openpyxl import load_workbook
from exlang import compile_xlang_to_xlsx


# ============================================================
# Roundtrip verification tests
# ============================================================

class TestRoundtripExamples:
    """Test that EXLANG semantics survive compilation to Excel."""

    def test_kpi_example(self, tmp_path):
        """KPI example from notebook compiles correctly."""
        xlang = """
        <xworkbook>
          <xsheet name="KPI">
            <xrow r="1"><xv>Region</xv><xv>Sales</xv></xrow>
            <xrow r="2"><xv>North</xv><xv>120000</xv></xrow>
            <xrow r="3"><xv>South</xv><xv>98000</xv></xrow>
            <xcell addr="A4" v="Total"/>
            <xcell addr="B4" v="=SUM(B2:B3)"/>
          </xsheet>
        </xworkbook>
        """
        output = tmp_path / "kpi.xlsx"
        compile_xlang_to_xlsx(xlang, output)

        wb = load_workbook(output, data_only=False)
        ws = wb["KPI"]

        # Verify structure
        assert ws["A1"].value == "Region"
        assert ws["B1"].value == "Sales"
        assert ws["A2"].value == "North"
        assert ws["B2"].value == 120000
        assert isinstance(ws["B2"].value, int)
        assert ws["A3"].value == "South"
        assert ws["B3"].value == 98000
        assert ws["A4"].value == "Total"
        assert ws["B4"].value == "=SUM(B2:B3)"

    def test_regional_sales_example(self, tmp_path):
        """Multi-sheet regional sales example."""
        xlang = """
        <xworkbook>
          <xsheet name="Data">
            <xrow r="1"><xv>Region</xv><xv>Q1</xv><xv>Q2</xv><xv>Q3</xv></xrow>
            <xrow r="2"><xv>North</xv><xv>120000</xv><xv>130000.5</xv><xv>125000</xv></xrow>
            <xrow r="3"><xv>South</xv><xv>95000</xv><xv>97000</xv><xv>99000</xv></xrow>
            <xrow r="4"><xv>East</xv><xv>88000</xv><xv>91000</xv><xv>-5000</xv></xrow>
            <xrow r="5"><xv>West</xv><xv>110000</xv><xv>115000</xv><xv>118000</xv></xrow>
          </xsheet>
          <xsheet name="Summary">
            <xrow r="1"><xv>Metric</xv><xv>Value</xv></xrow>
            <xrow r="2"><xv>TotalQ1</xv><xv>=SUM(Data!B2:B5)</xv></xrow>
            <xrow r="3"><xv>TotalAll</xv><xv>=SUM(Data!B2:D5)</xv></xrow>
            <xrow r="4"><xv>AverageQ2</xv><xv>=AVERAGE(Data!C2:C5)</xv></xrow>
          </xsheet>
        </xworkbook>
        """
        output = tmp_path / "regional_sales.xlsx"
        compile_xlang_to_xlsx(xlang, output)

        wb = load_workbook(output, data_only=False)

        # Verify Data sheet
        data_ws = wb["Data"]
        assert data_ws["A1"].value == "Region"
        assert data_ws["B2"].value == 120000
        assert isinstance(data_ws["B2"].value, int)
        assert data_ws["C2"].value == 130000.5
        assert isinstance(data_ws["C2"].value, float)
        assert data_ws["D4"].value == -5000

        # Verify Summary sheet
        summary_ws = wb["Summary"]
        assert summary_ws["A1"].value == "Metric"
        assert summary_ws["B2"].value == "=SUM(Data!B2:B5)"
        assert summary_ws["B3"].value == "=SUM(Data!B2:D5)"
        assert summary_ws["B4"].value == "=AVERAGE(Data!C2:C5)"

    def test_mixed_types_example(self, tmp_path):
        """Mixed types with non-default column starts."""
        xlang = """
        <xworkbook>
          <xsheet name="MixedTypes">
            <xrow r="5" c="B">
              <xv>ID</xv><xv>Code</xv><xv>Flag</xv><xv>Amount</xv>
            </xrow>
            <xrow r="6" c="B">
              <xv>1</xv><xv>00123</xv><xv>TRUE</xv><xv>1000.50</xv>
            </xrow>
            <xrow r="7" c="B">
              <xv>2</xv><xv>00456</xv><xv>FALSE</xv><xv>-250.75</xv>
            </xrow>
            <xrow r="8" c="B">
              <xv>3</xv><xv>00789</xv><xv>YES</xv><xv>500</xv>
            </xrow>
            <xcell addr="C6" v="00123" t="string"/>
            <xcell addr="C7" v="00456" t="string"/>
            <xcell addr="C8" v="00789" t="string"/>
            <xcell addr="D6" v="TRUE" t="bool"/>
            <xcell addr="D7" v="FALSE" t="bool"/>
            <xcell addr="D8" v="YES" t="bool"/>
            <xcell addr="B10" v="Total"/>
            <xcell addr="E10" v="=SUM(E6:E8)"/>
          </xsheet>
        </xworkbook>
        """
        output = tmp_path / "mixed_types.xlsx"
        compile_xlang_to_xlsx(xlang, output)

        wb = load_workbook(output, data_only=False)
        ws = wb["MixedTypes"]

        # Verify headers at row 5, starting column B
        assert ws["B5"].value == "ID"
        assert ws["C5"].value == "Code"
        assert ws["D5"].value == "Flag"
        assert ws["E5"].value == "Amount"

        # Verify data types
        assert ws["B6"].value == 1
        assert isinstance(ws["B6"].value, int)
        assert ws["C6"].value == "00123"  # Leading zeros preserved
        assert ws["D6"].value is True
        assert ws["E6"].value == 1000.50

        # Verify negative amount
        assert ws["E7"].value == -250.75

        # Verify total row
        assert ws["B10"].value == "Total"
        assert ws["E10"].value == "=SUM(E6:E8)"


# ============================================================
# Type preservation tests
# ============================================================

class TestTypePreservation:
    """Verify that data types are correctly preserved through compilation."""

    def test_numeric_type_preservation(self, tmp_path):
        """Integers remain int, floats remain float."""
        xlang = """
        <xworkbook>
          <xsheet name="Types">
            <xcell addr="A1" v="123"/>
            <xcell addr="A2" v="123.0"/>
            <xcell addr="A3" v="123.45"/>
          </xsheet>
        </xworkbook>
        """
        output = tmp_path / "numeric_types.xlsx"
        compile_xlang_to_xlsx(xlang, output)

        wb = load_workbook(output)
        ws = wb["Types"]

        assert isinstance(ws["A1"].value, int)
        assert ws["A1"].value == 123
        # Note: "123.0" is parsed as int 123 by our inference logic
        assert isinstance(ws["A2"].value, int)
        assert ws["A2"].value == 123
        assert isinstance(ws["A3"].value, float)
        assert ws["A3"].value == 123.45

    def test_formula_not_evaluated(self, tmp_path):
        """Formulas stored as strings, not evaluated."""
        xlang = """
        <xworkbook>
          <xsheet name="Formulas">
            <xcell addr="A1" v="10"/>
            <xcell addr="A2" v="20"/>
            <xcell addr="A3" v="=A1+A2"/>
          </xsheet>
        </xworkbook>
        """
        output = tmp_path / "formulas.xlsx"
        compile_xlang_to_xlsx(xlang, output)

        wb = load_workbook(output, data_only=False)
        ws = wb["Formulas"]

        # Formula should be stored, not evaluated
        assert ws["A3"].value == "=A1+A2"
        assert ws["A3"].data_type == "f"  # Formula type

    def test_string_preservation(self, tmp_path):
        """Strings with leading zeros, special chars preserved."""
        xlang = """
        <xworkbook>
          <xsheet name="Strings">
            <xcell addr="A1" v="00123" t="string"/>
            <xcell addr="A2" v="N/A"/>
            <xcell addr="A3" v="  spaces  "/>
          </xsheet>
        </xworkbook>
        """
        output = tmp_path / "strings.xlsx"
        compile_xlang_to_xlsx(xlang, output)

        wb = load_workbook(output)
        ws = wb["Strings"]

        assert ws["A1"].value == "00123"
        assert ws["A2"].value == "N/A"
        assert ws["A3"].value == "  spaces  "


# ============================================================
# Empty and edge case tests
# ============================================================

class TestEdgeCases:
    """Test edge cases and boundary conditions."""

    def test_empty_sheet(self, tmp_path):
        """Empty sheets are valid."""
        xlang = """
        <xworkbook>
          <xsheet name="Empty"></xsheet>
        </xworkbook>
        """
        output = tmp_path / "empty.xlsx"
        compile_xlang_to_xlsx(xlang, output)

        wb = load_workbook(output)
        assert "Empty" in wb.sheetnames

    def test_single_cell_workbook(self, tmp_path):
        """Minimal workbook with one cell."""
        xlang = """
        <xworkbook>
          <xsheet name="Single">
            <xcell addr="A1" v="Hello"/>
          </xsheet>
        </xworkbook>
        """
        output = tmp_path / "single.xlsx"
        compile_xlang_to_xlsx(xlang, output)

        wb = load_workbook(output)
        assert wb["Single"]["A1"].value == "Hello"

    def test_large_row_index(self, tmp_path):
        """Large row indices work correctly."""
        xlang = """
        <xworkbook>
          <xsheet name="Large">
            <xcell addr="A1000" v="Far down"/>
          </xsheet>
        </xworkbook>
        """
        output = tmp_path / "large.xlsx"
        compile_xlang_to_xlsx(xlang, output)

        wb = load_workbook(output)
        assert wb["Large"]["A1000"].value == "Far down"

    def test_large_column_index(self, tmp_path):
        """Large column indices work correctly."""
        xlang = """
        <xworkbook>
          <xsheet name="Wide">
            <xcell addr="ZZ1" v="Far right"/>
          </xsheet>
        </xworkbook>
        """
        output = tmp_path / "wide.xlsx"
        compile_xlang_to_xlsx(xlang, output)

        wb = load_workbook(output)
        assert wb["Wide"]["ZZ1"].value == "Far right"
