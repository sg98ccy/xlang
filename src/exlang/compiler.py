# ============================================================
# exlang.compiler: compile exlang to Excel
# ============================================================

from pathlib import Path
from xml.etree import ElementTree as ET
from jinja2 import Environment

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import range_boundaries

from .validator import validate_xlang_minimal
from .helpers import col_letter_to_index, infer_value, parse_range, parse_merge_range, substitute_template_vars


def preprocess_jinja_xlang(xlang_text: str) -> str:
    """
    Preprocess EXLang templates using Jinja2 with XML autoescape.
    
    This allows natural formula syntax:
        <xcell addr="A1" v="{{ formula }}"/>
    
    Where formula = '=IF(B1<100,"Low","High")' gets auto-escaped to:
        <xcell addr="A1" v="=IF(B1&lt;100,&quot;Low&quot;,&quot;High&quot;)"/>
    
    Uses Jinja2's industry-standard XML escaping mechanism.
    """
    # Create Jinja2 environment with XML autoescape enabled
    env = Environment(autoescape=True)
    
    # Render the template (this handles XML escaping automatically)
    template = env.from_string(xlang_text)
    
    # Render with empty context (no variables to substitute unless provided)
    return template.render()


def compile_xlang_to_xlsx(xlang_text: str, output_path: str | Path, **template_vars) -> None:
    """
    Compile a minimal subset of exlang into an Excel .xlsx file.
    
    Jinja2 preprocessing is ALWAYS enabled for automatic XML escaping and template support.
    This allows natural formula syntax without manual escaping.

    Supported tags:
      - xworkbook
      - xsheet
      - xrow
      - xv
      - xrepeat
      - xcell
      - xrange
      - xmerge
      - xstyle
    
    Args:
        xlang_text: EXLang XML string (with optional Jinja2 templates)
        output_path: Path to output .xlsx file
        **template_vars: Variables to pass to Jinja2 template rendering
    
    Example with template variables:
        xlang = '''
        <xworkbook>
          <xsheet name="Test">
            <xcell addr="A1" v="{{ formula }}"/>
          </xsheet>
        </xworkbook>
        '''
        compile_xlang_to_xlsx(xlang, "output.xlsx", 
                             formula='=IF(B1<100,"Low","High")')
    
    Example without template variables (still gets Jinja2 preprocessing):
        xlang = '''
        <xworkbook>
          <xsheet name="KPI">
            <xrow r="1"><xv>Region</xv><xv>Sales</xv></xrow>
          </xsheet>
        </xworkbook>
        '''
        compile_xlang_to_xlsx(xlang, "output.xlsx")
    """
    # ALWAYS preprocess with Jinja2 for automatic XML escaping
    if template_vars:
        xlang_text = Environment(autoescape=True).from_string(xlang_text).render(**template_vars)
    else:
        xlang_text = preprocess_jinja_xlang(xlang_text)
    
    root = ET.fromstring(xlang_text)

    errors = validate_xlang_minimal(root)
    if errors:
        formatted = "\n".join("  - " + e for e in errors)
        raise ValueError("Invalid XLang:\n" + formatted)

    wb = Workbook()
    wb.remove(wb.active)

    # Auto-generate sheet names for unnamed sheets
    auto_counter = 1
    for xsheet in root.findall("xsheet"):
        sheet_name = xsheet.attrib.get("name")
        if not sheet_name:
            sheet_name = f"Sheet{auto_counter}"
            auto_counter += 1
        ws = wb.create_sheet(title=sheet_name)

        # Process in order: xrow → xrange → xrepeat → xcell (last write wins)
        for xrow in xsheet.findall("xrow"):
            row_idx = int(xrow.attrib["r"])
            start_col_letter = xrow.attrib.get("c", "A")
            start_col_idx = col_letter_to_index(start_col_letter)

            for offset, xv in enumerate(xrow.findall("xv")):
                raw_value = xv.text or ""
                value = infer_value(raw_value, None)
                ws.cell(
                    row=row_idx,
                    column=start_col_idx + offset,
                    value=value,
                )

        for xrange in xsheet.findall("xrange"):
            from_addr = xrange.attrib["from"]
            to_addr = xrange.attrib["to"]
            fill_value = xrange.attrib["fill"]
            type_hint = xrange.attrib.get("t")
            
            from_row, from_col, to_row, to_col = parse_range(from_addr, to_addr)
            inferred_value = infer_value(fill_value, type_hint)
            
            for row in range(from_row, to_row + 1):
                for col in range(from_col, to_col + 1):
                    ws.cell(row=row, column=col, value=inferred_value)

        for xrepeat in xsheet.findall("xrepeat"):
            times = int(xrepeat.attrib["times"])
            direction = xrepeat.attrib.get("direction", "down")
            start_row = int(xrepeat.attrib.get("r", "1"))
            start_col_letter = xrepeat.attrib.get("c", "A")
            start_col_idx = col_letter_to_index(start_col_letter)
            
            # Extract template xv elements
            template_xvs = list(xrepeat.findall("xv"))
            
            for i in range(1, times + 1):
                # Calculate current position based on direction
                if direction == "down":
                    current_row = start_row + (i - 1)
                    current_col = start_col_idx
                else:  # direction == "right"
                    current_row = start_row
                    current_col = start_col_idx + (i - 1)
                
                # Process each xv in the template
                for offset, xv in enumerate(template_xvs):
                    raw_value = xv.text or ""
                    # Substitute template variables
                    substituted_value = substitute_template_vars(raw_value, i)
                    value = infer_value(substituted_value, None)
                    
                    # Calculate final cell position
                    if direction == "down":
                        ws.cell(row=current_row, column=current_col + offset, value=value)
                    else:  # direction == "right"
                        ws.cell(row=current_row + offset, column=current_col, value=value)

        for xcell in xsheet.findall("xcell"):
            addr = xcell.attrib["addr"]
            raw_value = xcell.attrib["v"]
            type_hint = xcell.attrib.get("t")
            value = infer_value(raw_value, type_hint)
            ws[addr] = value

        # Process xmerge (merge cells)
        for xmerge in xsheet.findall("xmerge"):
            addr = xmerge.attrib["addr"]
            # Parse merge range (e.g., "A1:B1")
            start_row, start_col, end_row, end_col = parse_merge_range(addr)
            ws.merge_cells(
                start_row=start_row,
                start_column=start_col,
                end_row=end_row,
                end_column=end_col
            )

        # Process xstyle (apply formatting)
        for xstyle in xsheet.findall("xstyle"):
            addr = xstyle.attrib["addr"]
            
            # Check if addr is a range or single cell
            if ":" in addr:
                # Range notation (e.g., "A1:B10")
                start_row, start_col, end_row, end_col = parse_merge_range(addr)
                cells_to_style = []
                for row in range(start_row, end_row + 1):
                    for col in range(start_col, end_col + 1):
                        cells_to_style.append(ws.cell(row=row, column=col))
            else:
                # Single cell
                cells_to_style = [ws[addr]]
            
            # Build Font object from attributes
            bold = xstyle.attrib.get("bold") == "true"
            italic = xstyle.attrib.get("italic") == "true"
            underline_val = "single" if xstyle.attrib.get("underline") == "true" else None
            
            font = Font(bold=bold, italic=italic, underline=underline_val)
            
            # Apply font to all cells
            for cell in cells_to_style:
                cell.font = font

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
